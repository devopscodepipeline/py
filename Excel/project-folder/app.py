from flask import Flask, render_template, request, send_file
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Font, Alignment
import os
from datetime import datetime

app = Flask(__name__)

# Directory to save output files
OUTPUT_DIR = "output_files"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Function to process files
def process_files(organized_file, unorganized_file):
    # Step 1: Load the organized file and extract merged headers
    wb = load_workbook(organized_file)
    ws = wb.active

    # Dynamically determine the range of columns
    max_col = ws.max_column
    merged_cells = []
    for merged in ws.merged_cells.ranges:
        merged_cells.append(merged)

    header_data = []
    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=max_col):
        header_row = [cell.value for cell in row]
        header_data.append(header_row)

    # Extract data headers from Row 4
    data_headers = [cell[0].value for cell in ws.iter_cols(min_row=4, max_row=4, min_col=1, max_col=max_col)]

    # Step 2: Load the unorganized file and filter matched columns
    df_unorganized = pd.read_excel(unorganized_file)

    # Match data headers with unorganized file columns
    matched_columns = [col for col in data_headers if col in df_unorganized.columns]
    matched_data = df_unorganized[matched_columns]
    matched_data = matched_data.reindex(columns=data_headers, fill_value=None)

    # Step 3: Create a new workbook for output
    wb_output = Workbook()
    ws_output = wb_output.active

    # Copy headers and merged cells
    for row_idx, row in enumerate(header_data, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws_output.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for merged in merged_cells:
        ws_output.merge_cells(str(merged))

    # Add Row 4 (data headers)
    for col_idx, header in enumerate(data_headers, start=1):
        cell = ws_output.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add matched data
    data_font = Font(size=10)
    for row_idx, row in enumerate(matched_data.itertuples(index=False), start=5):
        for col_idx, value in enumerate(row, start=1):
            cell = ws_output.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font

    # Apply borders dynamically
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    # Apply borders to all cells, including merged ranges
    for row in ws_output.iter_rows(min_row=1, max_row=ws_output.max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border

    for merged in ws_output.merged_cells.ranges:
        for row in ws_output.iter_rows(
            min_row=merged.min_row,
            max_row=merged.max_row,
            min_col=merged.min_col,
            max_col=merged.max_col
        ):
            for cell in row:
                cell.border = border

    # Generate a unique output file name
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(OUTPUT_DIR, f"output_{timestamp}.xlsx")
    wb_output.save(output_file)

    return output_file

# Flask route for the home page
@app.route('/')
def home():
    return render_template('index.html')

# Flask route to handle file upload and processing
@app.route('/process', methods=['POST'])
def process():
    if 'organized_file' not in request.files or 'unorganized_file' not in request.files:
        return "Please upload both files.", 400

    organized_file = request.files['organized_file']
    unorganized_file = request.files['unorganized_file']

    # Save uploaded files temporarily
    organized_file_path = os.path.join("temp", organized_file.filename)
    unorganized_file_path = os.path.join("temp", unorganized_file.filename)
    os.makedirs("temp", exist_ok=True)
    organized_file.save(organized_file_path)
    unorganized_file.save(unorganized_file_path)

    # Process files
    output_file = process_files(organized_file_path, unorganized_file_path)

    # Clean up temporary files
    os.remove(organized_file_path)
    os.remove(unorganized_file_path)

    # Send the output file to the user
    return send_file(output_file, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=True)
